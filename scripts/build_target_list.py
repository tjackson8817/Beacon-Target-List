"""
OT Cyber Consulting Target List Builder.

Given a seed company/website and/or NAICS codes and industry keywords,
builds an expanded target-company list (primary/secondary/tertiary) for
job-search networking, using only free/keyless public data plus the
Anthropic API (with web search enabled) for research assistance.

Data sources:
  - SEC EDGAR (browse-edgar by SIC code) for real, verifiable PUBLIC
    company data — free, no key.
  - Claude API with the web_search tool, for NAICS expansion, private-
    company discovery, and qualitative enrichment. This requires the
    same ANTHROPIC_API_KEY used elsewhere; it is not a new paid service.

What this deliberately does NOT do:
  - No LinkedIn scraping. "Target executives" comes back as SUGGESTED
    TITLES/keywords to search yourself, never as scraped named
    individuals or scraped contact URLs.
  - No fabricated precision. Revenue/employee data is only presented as
    verified when it comes from SEC EDGAR; otherwise it's labeled as an
    AI estimate requiring your own verification, exactly like the
    original workbook's own "Sources + Caveats" convention.
"""

import json
import os
import re
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from urllib.parse import quote

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "reports")
MODEL = "claude-sonnet-5"
API_KEY = os.environ.get("ANTHROPIC_API_KEY")

# SEC requires a descriptive User-Agent identifying the requester, and will
# block generic ones. Replace the email below with a real contact address.
SEC_HEADERS = {"User-Agent": "OT-Cyber-Target-List-Builder research-tool@example.com"}

# Starter NAICS -> SIC crosswalk for the codes most relevant to OT cyber
# consulting/advisory. Not exhaustive — codes outside this table fall back
# to asking Claude for the mapping.
NAICS_TO_SIC = {
    "541512": ["7379", "7371"],   # Computer Systems Design Services
    "541611": ["8742"],           # Admin/General Management Consulting
    "541618": ["8742"],           # Other Management Consulting
    "541690": ["8742", "8711"],   # Other Scientific/Technical Consulting
    "541330": ["8711"],           # Engineering Services
    "541519": ["7379"],           # Other Computer Related Services
    "541513": ["7379"],           # Computer Facilities Management
}

DEFAULT_SEED_NAICS = ["541512", "541611", "541618", "541690", "541330", "541519", "541513"]

SYSTEM_NAICS_EXPANSION = """You are an executive market research analyst helping a job seeker build a
target-company list. Given information about a seed company/industry, use
web search to confirm facts rather than relying on memory alone.

Return ONLY valid JSON, no preamble, no markdown fences, matching this schema:
{
  "primary_naics": [{"code": string, "description": string}],
  "secondary_naics": [{"code": string, "description": string, "relationship": string}],
  "rationale": string
}

"relationship" should briefly explain how this secondary code relates to the
primary business (e.g. "competitor", "supplier/vendor", "adjacent service
provider", "manufacturer served by this industry"). Do not invent NAICS codes
that don't exist — verify via search if uncertain."""

SYSTEM_DISCOVERY = """You are an executive market research analyst helping a job seeker identify
target companies for senior OT cybersecurity consulting/advisory roles. Use
web search actively; do not rely on memory alone for company names, and do
not fabricate companies that don't exist.

Return ONLY valid JSON: a list of objects, no preamble, no markdown fences:
[
  {"company": string, "website": string | null, "public_or_private": "public" | "private" | "unknown",
   "hq_state_or_country": string | null, "primary_naics_guess": string | null,
   "brief_reason": string}
]

Include a mix of company sizes where possible. Do not repeat any company in
the excluded list provided. If you cannot find enough real, verifiable
companies to reach the requested count, return fewer rather than inventing
placeholder companies."""

SYSTEM_ENRICHMENT = """You are an executive market research analyst enriching a target-company list
for a senior OT cybersecurity consulting/advisory job seeker. For each
company provided, use web search to check current, real information rather
than relying on memory alone.

Rules:
- Do not invent facts. If something can't be found, use null and say so via
  the confidence field rather than guessing silently.
- Revenue and employee count: if you are confident this is a public company
  with SEC filings, note that, but do NOT fabricate a specific dollar figure
  or headcount yourself — leave precise figures null (the pipeline fills
  verified figures from SEC EDGAR separately for public filers). For private
  companies, you may give a rough band (e.g. "$50M-$250M (estimate)") but
  MUST mark estimated fields with 'estimate' in the confidence field.
- "target_executive_titles" must be ROLE TITLES to search for (e.g. "VP
  Industrial Cybersecurity", "Director OT Security Practice"), never named
  individuals — this pipeline does not look up or report named executives.
- "linkedin_search_keywords" should be a short list of search terms/phrases
  useful for manually searching LinkedIn yourself.
- "currently_hiring_signal" should reflect what you can find via web search
  right now (careers page activity, recent job postings relevant to OT/
  cyber roles) — this is a fast-changing fact, so treat it as a snapshot at
  research time, not a stable attribute like HQ location.
- severity/priority fields should be your qualitative assessment, clearly
  that -- an assessment, not verified fact.

Return ONLY valid JSON: a list of objects matching this schema, no preamble,
no markdown fences:
[
  {
    "company": string,
    "website": string | null,
    "category": string,
    "hq": string | null,
    "industry_emphasis": string,
    "company_size_band": string | null,
    "revenue_band": string | null,
    "confidence": "public-company (see EDGAR)" | "estimate" | "unknown",
    "ot_ics_relevance": "High" | "Medium" | "Low",
    "cybersecurity_relevance": "High" | "Medium" | "Low",
    "why_this_company_fits": string,
    "growth_signals": string,
    "outreach_strategy": string,
    "target_executive_titles": [string],
    "linkedin_search_keywords": [string],
    "currently_hiring_signal": "Yes" | "No" | "Unclear",
    "currently_hiring_detail": string | null,
    "priority_score": number,
    "source_url": string | null
  }
]"""


def call_claude(system, user_content, max_tokens=3000, use_web_search=True):
    body = {
        "model": MODEL,
        "max_tokens": max_tokens,
        "system": system,
        "messages": [{"role": "user", "content": user_content}],
    }
    if use_web_search:
        body["tools"] = [{"type": "web_search_20250305", "name": "web_search"}]

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": API_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json=body,
        timeout=120,
    )
    resp.raise_for_status()
    data = resp.json()
    text = "".join(block.get("text", "") for block in data.get("content", []))
    text = re.sub(r"^```json|^```|```$", "", text.strip(), flags=re.MULTILINE).strip()

    try:
        return json.loads(text)
    except json.JSONDecodeError as exc:
        for open_c, close_c in [("{", "}"), ("[", "]")]:
            start, end = text.find(open_c), text.rfind(close_c)
            if start != -1 and end != -1 and end > start:
                try:
                    return json.loads(text[start:end + 1])
                except json.JSONDecodeError:
                    continue
        snippet = text[:400].replace("\n", " ")
        raise ValueError(f"{exc} -- raw response started with: {snippet!r}") from exc


def naics_to_sic(code):
    if code in NAICS_TO_SIC:
        return NAICS_TO_SIC[code]
    try:
        result = call_claude(
            "Return ONLY a JSON array of SIC code strings (e.g. [\"7379\"]) that correspond to the "
            "given NAICS code, based on official Census/BLS crosswalks. If uncertain, verify via search. "
            "No preamble, no markdown fences.",
            f"NAICS code: {code}",
            max_tokens=300,
        )
        return result if isinstance(result, list) else []
    except Exception as exc:
        print(f"SIC lookup failed for NAICS {code}: {exc}", file=sys.stderr)
        return []


def edgar_companies_by_sic(sic_code, limit=20):
    """Real, free, verifiable public-company discovery via SEC EDGAR."""
    companies = []
    try:
        resp = requests.get(
            "https://www.sec.gov/cgi-bin/browse-edgar",
            params={
                "action": "getcompany", "SIC": sic_code, "type": "10-K",
                "dateb": "", "owner": "include", "count": limit, "output": "atom",
            },
            headers=SEC_HEADERS, timeout=20,
        )
        resp.raise_for_status()
        root = ET.fromstring(resp.content)
        ns = {"atom": "http://www.w3.org/2005/Atom"}
        for entry in root.findall(".//atom:entry", ns):
            title_el = entry.find("atom:title", ns)
            title = title_el.text.strip() if title_el is not None and title_el.text else ""
            if not title:
                continue
            # Titles are typically "CIK - Company Name" or just "Company Name"
            match = re.match(r"^(\d{4,10})\s*-\s*(.+)$", title)
            if match:
                cik, name = match.group(1), match.group(2).strip()
            else:
                cik, name = None, title
            companies.append({"company": name, "cik": cik, "sic": sic_code, "source": "SEC EDGAR"})
    except (requests.RequestException, ET.ParseError) as exc:
        print(f"EDGAR lookup failed for SIC {sic_code}: {exc}", file=sys.stderr)
    time.sleep(0.15)  # stay comfortably under SEC's 10 req/sec limit
    return companies


def dedupe_companies(companies):
    seen = set()
    out = []
    for c in companies:
        key = re.sub(r"[^a-z0-9]", "", c["company"].lower())
        if key and key not in seen:
            seen.add(key)
            out.append(c)
    return out


def expand_naics(seed_company, seed_website, seed_naics, industry_keywords):
    if seed_naics:
        primary = [{"code": c.strip(), "description": ""} for c in seed_naics.split(",") if c.strip()]
        user_content = (
            f"Seed company: {seed_company or '(not specified)'}\n"
            f"Seed website: {seed_website or '(not specified)'}\n"
            f"Given primary NAICS codes: {seed_naics}\n"
            f"Industry keywords: {industry_keywords or '(none)'}\n\n"
            f"Confirm/describe these primary codes and suggest secondary/tertiary related NAICS codes "
            f"(competitors, suppliers, manufacturers served, adjacent service providers) relevant for an "
            f"OT cybersecurity consulting/advisory job search."
        )
    else:
        user_content = (
            f"Seed company: {seed_company or '(not specified)'}\n"
            f"Seed website: {seed_website or '(not specified)'}\n"
            f"Industry keywords: {industry_keywords or '(none)'}\n\n"
            f"Identify the likely primary NAICS code(s) for this company/industry, and suggest "
            f"secondary/tertiary related NAICS codes (competitors, suppliers, manufacturers served, "
            f"adjacent service providers) relevant for an OT cybersecurity consulting/advisory job search."
        )
    return call_claude(SYSTEM_NAICS_EXPANSION, user_content, max_tokens=2000)


def discover_via_ai(naics_codes, keywords, geography, exclude_names, count):
    user_content = (
        f"NAICS codes: {', '.join(naics_codes)}\n"
        f"Keywords: {keywords or '(none)'}\n"
        f"Geography: {geography or 'United States'}\n"
        f"Already found (exclude these): {', '.join(exclude_names) if exclude_names else '(none)'}\n\n"
        f"Identify up to {count} additional real companies matching these criteria."
    )
    result = call_claude(SYSTEM_DISCOVERY, user_content, max_tokens=3000)
    return result if isinstance(result, list) else []


def enrich_batch(companies_batch, context_naics, context_keywords):
    names = [c["company"] for c in companies_batch]
    user_content = (
        f"Context — relevant NAICS codes: {', '.join(context_naics)}. Keywords: {context_keywords or '(none)'}.\n\n"
        f"Enrich these companies: {', '.join(names)}"
    )
    result = call_claude(SYSTEM_ENRICHMENT, user_content, max_tokens=4000)
    return result if isinstance(result, list) else []


def edgar_company_facts(cik):
    """Pull verified public-filer facts (SIC, name, address) for a known CIK."""
    if not cik:
        return {}
    try:
        padded = cik.zfill(10)
        resp = requests.get(
            f"https://data.sec.gov/submissions/CIK{padded}.json",
            headers=SEC_HEADERS, timeout=20,
        )
        resp.raise_for_status()
        data = resp.json()
        addr = data.get("addresses", {}).get("business", {})
        hq = ", ".join(filter(None, [addr.get("city"), addr.get("stateOrCountry")]))
        return {"sic": data.get("sicDescription"), "hq_verified": hq or None, "name_verified": data.get("name")}
    except requests.RequestException:
        return {}
    finally:
        time.sleep(0.15)


def slugify(name):
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-") or "unnamed-list"


FILL_HEADER = PatternFill("solid", start_color="FF1F4B5C", end_color="FF1F4B5C")
FILL_HIGH = PatternFill("solid", start_color="FFE8F2EA", end_color="FFE8F2EA")
FILL_MEDIUM = PatternFill("solid", start_color="FFFAF0E0", end_color="FFFAF0E0")
FILL_LOW = PatternFill("solid", start_color="FFECEEEC", end_color="FFECEEEC")
FONT_HIGH = Font(color="FF2E6B46")
FONT_MEDIUM = Font(color="FFA4650A")
FONT_LOW = Font(color="FF5F6864")

COMPANY_COLUMNS = [
    ("Company", "company"), ("Website", "website"), ("Category", "category"), ("HQ", "hq"),
    ("Industry Emphasis", "industry_emphasis"), ("Company Size", "company_size_band"),
    ("Revenue Band", "revenue_band"), ("Confidence", "confidence"),
    ("OT/ICS Relevance", "ot_ics_relevance"), ("Cybersecurity Relevance", "cybersecurity_relevance"),
    ("Why This Company Fits", "why_this_company_fits"), ("Growth Signals", "growth_signals"),
    ("Outreach Strategy", "outreach_strategy"), ("Target Executive Titles", "_titles"),
    ("LinkedIn Search Keywords", "_keywords"), ("Currently Hiring", "currently_hiring_signal"),
    ("Hiring Detail", "currently_hiring_detail"), ("Priority Score", "priority_score"),
    ("Discovery Source", "discovery_source"), ("Source URL", "source_url"),
    ("Status", "_status"), ("Contact Name", "_blank"), ("Notes", "_blank"),
]

RELEVANCE_STYLE = {
    "high": (FILL_HIGH, FONT_HIGH), "medium": (FILL_MEDIUM, FONT_MEDIUM), "low": (FILL_LOW, FONT_LOW)
}


def style_header_row(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = FILL_HEADER
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def write_xlsx(report, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Target Companies"

    for col, (label, _) in enumerate(COMPANY_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=label)
    style_header_row(ws, len(COMPANY_COLUMNS))

    for r, c in enumerate(report.get("companies", []), start=2):
        for col, (label, key) in enumerate(COMPANY_COLUMNS, start=1):
            if key == "_titles":
                value = "; ".join(c.get("target_executive_titles") or [])
            elif key == "_keywords":
                value = "; ".join(c.get("linkedin_search_keywords") or [])
            elif key == "_status":
                value = "Not Started"
            elif key == "_blank":
                value = ""
            else:
                value = c.get(key)
            cell = ws.cell(row=r, column=col, value=value if value is not None else "")

            if key in ("ot_ics_relevance", "cybersecurity_relevance"):
                style = RELEVANCE_STYLE.get((value or "").lower())
                if style:
                    cell.fill, cell.font = style
            elif key == "confidence":
                if value and str(value).startswith("public-company"):
                    cell.fill, cell.font = FILL_HIGH, FONT_HIGH
                elif value == "estimate":
                    cell.fill, cell.font = FILL_MEDIUM, FONT_MEDIUM
                else:
                    cell.fill, cell.font = FILL_LOW, FONT_LOW

    widths = [22, 20, 18, 20, 26, 14, 14, 16, 12, 14, 32, 26, 26, 26, 26, 12, 24, 10, 16, 30, 12, 16, 20]
    for i, w in enumerate(widths[:len(COMPANY_COLUMNS)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # NAICS Codes sheet
    ws2 = wb.create_sheet("NAICS Codes")
    ws2.append(["NAICS Code", "Description", "Type", "Relationship"])
    style_header_row(ws2, 4)
    for n in report.get("primary_naics", []):
        ws2.append([n.get("code"), n.get("description", ""), "Primary", ""])
    for n in report.get("secondary_naics", []):
        ws2.append([n.get("code"), n.get("description", ""), "Secondary", n.get("relationship", "")])
    for i, w in enumerate([16, 44, 12, 24], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Search Info sheet
    ws3 = wb.create_sheet("Search Info")
    info_rows = [
        ("List Name", report.get("list_name")), ("Seed Company", report.get("seed_company")),
        ("Seed Website", report.get("seed_website")), ("Industry Keywords", report.get("industry_keywords")),
        ("Geography", report.get("geography")), ("Generated", report.get("generated_date")),
        ("NAICS Rationale", report.get("naics_rationale")),
    ]
    ws3.append(["Field", "Value"])
    style_header_row(ws3, 2)
    for label, value in info_rows:
        ws3.append([label, value or ""])
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 70
    for row in ws3.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(path)


def update_manifest(report, slug):
    manifest_path = os.path.join(DATA_DIR, "manifest.json")
    manifest = []
    if os.path.exists(manifest_path):
        try:
            with open(manifest_path) as f:
                manifest = json.load(f)
        except json.JSONDecodeError:
            manifest = []
    manifest = [m for m in manifest if m.get("slug") != slug]
    manifest.append({
        "slug": slug, "list_name": report["list_name"], "seed_company": report["seed_company"],
        "generated_date": report["generated_date"], "total_companies": len(report["companies"])
    })
    manifest.sort(key=lambda m: m["generated_date"], reverse=True)
    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2)


def main():
    if not API_KEY:
        print("ANTHROPIC_API_KEY is not set.", file=sys.stderr)
        sys.exit(1)

    list_name = os.environ.get("LIST_NAME", "").strip() or "Untitled Target List"
    seed_company = os.environ.get("SEED_COMPANY", "").strip()
    seed_website = os.environ.get("SEED_WEBSITE", "").strip()
    seed_naics = os.environ.get("SEED_NAICS", "").strip()
    industry_keywords = os.environ.get("INDUSTRY_KEYWORDS", "").strip()
    geography = os.environ.get("GEOGRAPHY", "").strip()
    max_companies = int(os.environ.get("MAX_COMPANIES", "25") or 25)
    skip_ai_enrichment = os.environ.get("SKIP_AI_ENRICHMENT", "").strip().lower() in ("true", "1", "yes")

    print(f"Building target list: {list_name}")

    print("Step 1: NAICS expansion...")
    naics_result = expand_naics(seed_company, seed_website, seed_naics, industry_keywords)
    primary_codes = [c["code"] for c in naics_result.get("primary_naics", [])] or DEFAULT_SEED_NAICS
    secondary_codes = [c["code"] for c in naics_result.get("secondary_naics", [])]
    all_codes = list(dict.fromkeys(primary_codes + secondary_codes))
    print(f"  Primary: {primary_codes}")
    print(f"  Secondary: {secondary_codes}")

    print("Step 2: SEC EDGAR public-company discovery...")
    edgar_found = []
    for code in all_codes:
        for sic in naics_to_sic(code):
            found = edgar_companies_by_sic(sic, limit=max(5, max_companies // len(all_codes)))
            print(f"  NAICS {code} -> SIC {sic}: {len(found)} compan(y/ies)")
            edgar_found.extend(found)
    edgar_found = dedupe_companies(edgar_found)[:max_companies]

    if skip_ai_enrichment:
        print("Quick mode: skipping AI discovery and enrichment (SIGNAL_AI_ENRICHMENT=false).")
        print("Only free SEC EDGAR data will be included — expect mostly public companies, unenriched.")
        enriched = []
        for c in edgar_found:
            verified = edgar_company_facts(c.get("cik")) if c.get("cik") else {}
            enriched.append({
                "company": c["company"], "website": None, "category": None,
                "hq": verified.get("hq_verified"), "industry_emphasis": None,
                "company_size_band": None, "revenue_band": None,
                "confidence": "public-company (see EDGAR)" if verified.get("sic") else "unknown",
                "ot_ics_relevance": None, "cybersecurity_relevance": None,
                "why_this_company_fits": "Quick mode — run full enrichment to populate this.",
                "growth_signals": None, "outreach_strategy": None,
                "target_executive_titles": [], "linkedin_search_keywords": [],
                "currently_hiring_signal": None, "currently_hiring_detail": None,
                "priority_score": None, "source_url": None,
                "sic_verified": verified.get("sic"), "discovery_source": c.get("source", "SEC EDGAR"),
            })
    else:
        remaining_slots = max(0, max_companies - len(edgar_found))
        ai_found = []
        if remaining_slots > 0:
            print(f"Step 3: AI-assisted discovery for {remaining_slots} more companies...")
            exclude = [c["company"] for c in edgar_found]
            ai_found = discover_via_ai(all_codes, industry_keywords, geography, exclude, remaining_slots)
            ai_found = [{"company": c.get("company", ""), "cik": None, "source": "AI research (verify)",
                         **c} for c in ai_found if c.get("company")]

        combined = dedupe_companies(edgar_found + ai_found)[:max_companies]
        print(f"Total companies to enrich: {len(combined)}")

        print("Step 4: Enriching companies (batched)...")
        enriched = []
        batch_size = 5
        for i in range(0, len(combined), batch_size):
            batch = combined[i:i + batch_size]
            print(f"  Enriching batch {i // batch_size + 1} ({len(batch)} companies)...")
            try:
                batch_result = enrich_batch(batch, all_codes, industry_keywords)
            except Exception as exc:
                print(f"  Batch failed: {exc}", file=sys.stderr)
                continue
            for item in batch_result:
                source_match = next((c for c in batch if c["company"].lower() in item.get("company", "").lower()
                                      or item.get("company", "").lower() in c["company"].lower()), {})
                if source_match.get("cik"):
                    verified = edgar_company_facts(source_match["cik"])
                    if verified.get("sic"):
                        item["confidence"] = "public-company (see EDGAR)"
                        item["hq"] = verified.get("hq_verified") or item.get("hq")
                        item["sic_verified"] = verified.get("sic")
                item["discovery_source"] = source_match.get("source", "unknown")
                enriched.append(item)

    report = {
        "list_name": list_name,
        "seed_company": seed_company or "(not specified)",
        "seed_website": seed_website or None,
        "industry_keywords": industry_keywords or None,
        "geography": geography or None,
        "primary_naics": naics_result.get("primary_naics", []),
        "secondary_naics": naics_result.get("secondary_naics", []),
        "naics_rationale": naics_result.get("rationale", ""),
        "generated_date": datetime.now(timezone.utc).isoformat(),
        "companies": enriched,
    }

    os.makedirs(DATA_DIR, exist_ok=True)
    slug = slugify(list_name)
    with open(os.path.join(DATA_DIR, f"{slug}.json"), "w") as f:
        json.dump(report, f, indent=2)
    write_xlsx(report, os.path.join(DATA_DIR, f"{slug}.xlsx"))
    update_manifest(report, slug)

    print(f"\nDone. {len(enriched)} companies written to data/reports/{slug}.json and {slug}.xlsx")


if __name__ == "__main__":
    main()
